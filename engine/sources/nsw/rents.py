"""NSW weekly rents from Fair Trading's rental bond lodgement workbooks.

The CKAN entries only link the landing page, so we scrape it for the monthly
"RentalBond_Lodgements_<Month>_<Year>.xlsx" files and read the latest 12.
Each workbook lists individual lodgements with postcode, dwelling type,
bedrooms and weekly rent; medians per postcode flow to SA2s through the
schools-derived suburb/postcode pairs. rent_12m ships as None in v1 (it
would need another year of workbooks).
"""
from __future__ import annotations

import re
import statistics
from collections import defaultdict

import openpyxl
import requests

from ... import config
from ...fetch import fresh
from ..crime import _sa2_localities
import json

PAGE = "https://www.fairtrading.nsw.gov.au/about-fair-trading/rental-bond-data"
_MONTHS = {m: i + 1 for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"])}

# The Fair Trading site serves the landing page to anyone but gates the xlsx
# asset URLs behind a JS challenge — run #4 (bot UA) and run #5 (browser UA +
# landing-page cookies) both got an HTML page back for every workbook. So like
# the Valuer General zips, direct download is attempted first but the Wayback
# Machine is the fallback that actually works.
_SESSION = requests.Session()
_SESSION.headers.update({
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
    "Referer": PAGE,
})


def _fetch_xlsx(url: str, filename: str):
    """Download a bond workbook, validating it's a real zip (xlsx) not an HTML page.

    Also self-heals cache poisoning: earlier runs stored challenge HTML under
    .xlsx names, and those bad files live on in the Actions data_raw cache.
    """
    from ...fetch import fetch_wayback
    config.DATA_RAW.mkdir(parents=True, exist_ok=True)
    dest = config.DATA_RAW / filename
    if dest.exists() and dest.stat().st_size > 0:
        if dest.read_bytes()[:2] == b"PK":
            print(f"  cached  {filename} ({dest.stat().st_size/1e6:.1f} MB)")
            return dest
        print(f"  rents: cached {filename} is not a zip — re-downloading")
        dest.unlink()
    try:
        r = _SESSION.get(url, timeout=120)
        r.raise_for_status()
        if r.content[:2] != b"PK":
            raise RuntimeError(f"got {r.headers.get('content-type')!r} instead of xlsx")
        dest.write_bytes(r.content)
        print(f"  downloaded  {filename} ({dest.stat().st_size/1e6:.1f} MB)")
        return dest
    except Exception as e:  # noqa: BLE001 - JS-challenged asset URL: go via Wayback
        print(f"  rents: direct download blocked ({e}) — trying Wayback")
    path = fetch_wayback(url, filename)
    if path.read_bytes()[:2] != b"PK":
        path.unlink()
        raise RuntimeError("Wayback copy is not an xlsx either")
    return path


def _lodgement_urls() -> list[tuple[int, int, str]]:
    """[(year, month, absolute url)] for every lodgement workbook on the page."""
    html = _SESSION.get(PAGE, timeout=60).text
    out = []
    for href in set(re.findall(r'href="([^"]+\.xlsx[^"]*)"', html, re.I)):
        m = re.search(r"odgements?[_-]([A-Za-z]+)[_-](\d{4})", href)
        if not m:
            continue
        mon = _MONTHS.get(m.group(1).lower())
        if not mon:
            continue
        url = href if href.startswith("http") else "https://www.fairtrading.nsw.gov.au" + href
        out.append((int(m.group(2)), mon, url))
    return sorted(out)


def _rents_by_postcode() -> tuple[dict[str, dict], str]:
    """({postcode: {"all": [...], "house3": [...], "flat2": [...], "n": int}}, latest label)"""
    cache = config.DATA_RAW / "nsw_bond_rents.json"
    if fresh(cache, 45):
        d = json.loads(cache.read_text(encoding="utf-8"))
        if d.get("postcodes"):   # an empty cache means a failed run — rebuild
            print("  cached  nsw_bond_rents.json")
            return d["postcodes"], d["latest"]

    urls = _lodgement_urls()
    if not urls:
        raise RuntimeError("no lodgement workbooks found on the Fair Trading page")
    # newest first, walking back past months Wayback hasn't archived, until 12 parse
    label = None
    parsed = 0
    acc: dict[str, dict] = defaultdict(lambda: {"all": [], "house3": [], "flat2": [], "n": 0})
    for year, mon, url in reversed(urls[-24:]):
        if parsed >= 12:
            break
        try:
            path = _fetch_xlsx(url, f"nsw_bonds_{year}_{mon:02d}.xlsx")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header = None
            for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                cells = [str(c or "").strip().lower() for c in row]
                if any("postcode" in c for c in cells) and any("rent" in c for c in cells):
                    header = cells
                    break
            if header is None:
                print(f"  rents: no header row in {url.rsplit('/', 1)[-1]} — skipped")
                continue
            i_pc = next(i for i, c in enumerate(header) if "postcode" in c)
            i_rent = next(i for i, c in enumerate(header) if "rent" in c)
            i_dw = next((i for i, c in enumerate(header) if "dwelling" in c), None)
            i_bed = next((i for i, c in enumerate(header) if "bedroom" in c), None)
            started = False
            for row in ws.iter_rows(values_only=True):
                cells = [str(c or "").strip().lower() for c in row]
                if not started:
                    started = cells == header
                    continue
                pc = str(row[i_pc] or "").strip().split(".")[0]
                try:
                    rent = float(row[i_rent])
                except (TypeError, ValueError):
                    continue
                if len(pc) != 4 or not pc.isdigit() or not (30 <= rent <= 20000):
                    continue
                a = acc[pc]
                a["all"].append(rent)
                a["n"] += 1
                dw = str(row[i_dw] or "").lower() if i_dw is not None else ""
                bed = str(row[i_bed] or "").strip().split(".")[0] if i_bed is not None else ""
                if dw.startswith("h") and bed == "3":
                    a["house3"].append(rent)
                elif dw.startswith(("f", "u")) and bed == "2":
                    a["flat2"].append(rent)
            wb.close()
            parsed += 1
            label = label or f"{year}-{mon:02d}"
        except Exception as e:  # noqa: BLE001 - one bad month shouldn't kill rents
            print(f"  rents: {url.rsplit('/', 1)[-1]} failed ({e})")
    label = label or f"{urls[-1][0]}-{urls[-1][1]:02d}"
    out = {pc: a for pc, a in acc.items() if a["n"] >= 5}
    if out:   # never cache a failed (empty) run
        cache.write_text(json.dumps({"postcodes": out, "latest": label}), encoding="utf-8")
    print(f"  rents: bond medians for {len(out)} NSW postcodes (12 months to {label})")
    return out, label


def get_rents(name_by_code: dict[str, str], lga_by_code: dict[str, str]) -> dict[str, dict]:
    from .schools import suburb_postcode_pairs
    by_pc, label = _rents_by_postcode()

    loc2pcs: dict[str, set] = defaultdict(set)
    for loc, pc in suburb_postcode_pairs():
        loc2pcs[loc].add(pc)

    out = {}
    for code, name in name_by_code.items():
        pcs = set()
        for loc in _sa2_localities(name):
            pcs |= loc2pcs.get(loc, set())
        allr, h3, f2, n = [], [], [], 0
        for pc in pcs:
            a = by_pc.get(pc)
            if not a:
                continue
            allr += a["all"]; h3 += a["house3"]; f2 += a["flat2"]; n += a["n"]
        if len(allr) < 5:
            out[code] = {}
            continue
        out[code] = {
            "rent_weekly": round(statistics.median(allr)),
            "rent_12m": None,
            "rent_bonds": n,
            "rent_quarter": label,
            "rent_source": "postcode",
            "house_rent": round(statistics.median(h3)) if len(h3) >= 5 else None,
            "flat_rent": round(statistics.median(f2)) if len(f2) >= 5 else None,
        }
    matched = sum(1 for v in out.values() if v)
    print(f"  rents: weekly medians for {matched}/{len(out)} SA2s (bond lodgements)")
    return out
