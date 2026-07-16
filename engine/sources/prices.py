"""Property-price inputs per SA2 from the Victorian Valuer-General.

Source: Victorian Property Sales Report — "Median <House|Unit> by Suburb Time
Series" (Land.Vic / Valuer-General), 10-year annual median sale prices by
suburb. Official, free, and the most reliable bulk source for Melbourne.

We compute, per suburb: latest median, 12-month % change, and 3-year CAGR, then
attach to SA2s by suburb-name matching (an SA2 name like "Tarneit - North" or
"Carlton North - Princes Hill" maps to one or more localities; we average).
Price varies far less within a suburb than crime does within an LGA, so a
suburb median applied to its child SA2s is a sound approximation.

land.vic.gov.au is behind a bot WAF, so the files are pulled via the Wayback
Machine (see engine/fetch.fetch_wayback). Both houses and units now use the
2014-2024 series (the 2014-2024 unit file was archived Nov 2025).

Also exports the full yearly house-median series per SA2 for the scorecard
sparkline.
"""
from __future__ import annotations

import re

import openpyxl

from ..fetch import fetch_wayback

HOUSES_URL = "https://www.land.vic.gov.au/__data/assets/excel_doc/0032/756581/houses-by-suburb-2014-2024.xlsx"
UNITS_URL = "https://www.land.vic.gov.au/__data/assets/excel_doc/0033/756582/units-by-suburb-2014-2024.xlsx"


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("^", "").replace("$", "").replace(",", "").strip()
    if s in ("", "-", "NA", "N/A", "n/a"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _load_series(path) -> dict[str, dict[int, float]]:
    """{LOCALITY: {year: median}} from a VG 'by suburb time series' workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    # Locate the header row that lists the years and map year -> column index.
    year_col, header_i = {}, None
    for ri, row in enumerate(rows[:8]):
        ys = {v: ci for ci, v in enumerate(row) if isinstance(v, int) and 2010 <= v <= 2026}
        if len(ys) >= 5:
            year_col, header_i = ys, ri
            break
    data = {}
    for row in rows[header_i + 1:]:
        loc = row[0]
        if not isinstance(loc, str):
            continue
        loc = loc.strip().upper()
        if not loc or loc == "LOCALITY":
            continue
        years = {y: _num(row[ci]) for y, ci in year_col.items() if ci < len(row)}
        if any(v for v in years.values()):
            data[loc] = years
    return data


def _metrics(years: dict[int, float]) -> dict:
    """latest median + 12m % change + 3yr CAGR % from a year->median series."""
    have = sorted(y for y, v in years.items() if v)
    if not have:
        return {}
    latest = have[-1]
    med = years[latest]
    out = {"median": med, "year": latest, "m12": None, "cagr3": None}
    prev = years.get(latest - 1)
    if prev:
        out["m12"] = round((med - prev) / prev * 100, 1)
    y3 = years.get(latest - 3)
    if y3 and y3 > 0:
        out["cagr3"] = round(((med / y3) ** (1 / 3) - 1) * 100, 1)
    return out


def _candidates(name: str) -> list[str]:
    """Suburb localities a given SA2 name could correspond to (UPPERCASE)."""
    n = re.sub(r"\(.*?\)", "", name)                  # drop "(Vic.)" etc.
    parts = [p.strip() for p in re.split(r"\s*-\s*", n) if p.strip()]
    cands = [p.upper() for p in parts]
    cands.append(re.sub(r"\s+", " ", n).strip().upper())  # whole name too
    seen, out = set(), []
    for c in cands:
        if c and c not in seen:
            seen.add(c); out.append(c)
    return out


def _avg(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else None


def get_prices(name_by_code: dict[str, str]) -> dict[str, dict]:
    """{sa2_code: {median_house, median_unit, house_12m, house_3yr_cagr, matched}}"""
    houses = _load_series(fetch_wayback(HOUSES_URL, "vg_houses_by_suburb.xlsx"))
    units = _load_series(fetch_wayback(UNITS_URL, "vg_units_by_suburb.xlsx"))
    hm = {loc: _metrics(y) for loc, y in houses.items()}
    um = {loc: _metrics(y) for loc, y in units.items()}

    out = {}
    for code, name in name_by_code.items():
        cands = _candidates(name)
        h = [hm[c] for c in cands if c in hm and hm[c]]
        u = [um[c] for c in cands if c in um and um[c]]
        matched = [c for c in cands if c in hm]
        # yearly house-median series (averaged across matched localities) for sparklines
        series = {}
        for c in matched:
            for y, v in houses[c].items():
                if v:
                    series.setdefault(y, []).append(v)
        house_series = [[y, round(sum(vs) / len(vs))] for y, vs in sorted(series.items())]
        out[code] = {
            "median_house": _avg([m["median"] for m in h]),
            "house_12m": _avg([m["m12"] for m in h]),
            "house_3yr_cagr": _avg([m["cagr3"] for m in h]),
            "house_year": max((m["year"] for m in h), default=None),
            "median_unit": _avg([m["median"] for m in u]),
            "unit_12m": _avg([m["m12"] for m in u]),
            "unit_3yr_cagr": _avg([m["cagr3"] for m in u]),
            "unit_year": max((m["year"] for m in u), default=None),
            "house_series": house_series,
            "matched": matched,
        }
    cov = sum(1 for v in out.values() if v["median_house"])
    print(f"  prices: matched house medians for {cov}/{len(out)} SA2s")
    return out


if __name__ == "__main__":  # pragma: no cover
    import json
    from .. import config
    fc = json.loads((config.CITY_DATA / config.BOUNDARIES_NAME).read_text(encoding="utf-8"))
    names = {f["properties"]["sa2_code"]: f["properties"]["sa2_name"] for f in fc["features"]}
    pr = get_prices(names)
    for nm in ("Nunawading", "Toorak", "Brighton (Vic.)", "Tarneit - North", "Cobblebank - Strathtulloh"):
        code = next((c for c, n in names.items() if n == nm), None)
        if code:
            print(f"  {nm:28} {pr[code]}")
