"""Estimated Resident Population (ERP) per SA2 — ABS Regional Population.

Census 2021 counts are badly stale in growth corridors (Tarneit-North has
roughly doubled), which inflated per-100k crime rates exactly where the Invest
lens points. This module pulls the ABS "Regional population" data cube
(32180DS0001, ERP at 30 June by SA2) and returns the latest year's estimate.

Population *structure* ratios (child share, tenure) stay on Census 2021 — the
mix moves far more slowly than the level.
"""
from __future__ import annotations

import json

import openpyxl

from .. import config
from ..fetch import fetch

ERP_URL = ("https://www.abs.gov.au/statistics/people/population/regional-population/"
           "2024-25/32180DS0001_2024-25.xlsx")
# Next release's URL, tried first so the data self-upgrades when ABS publishes
# (typically each March); falls back to the pinned URL above.
ERP_NEXT_URL = ("https://www.abs.gov.au/statistics/people/population/regional-population/"
                "2025-26/32180DS0001_2025-26.xlsx")


def get_erp() -> dict[str, dict]:
    """{sa2_code: {"erp": latest ERP, "erp_year": June year}} for Victorian SA2s."""
    from ..fetch import fresh
    cache = config.DATA_RAW / "abs_erp_sa2.json"
    if fresh(cache, 90):
        print("  cached  abs_erp_sa2.json")
        return json.loads(cache.read_text(encoding="utf-8"))

    try:
        path = fetch(ERP_NEXT_URL, "abs_erp_sa2.xlsx", max_age_days=90)
    except Exception:
        print("  erp: next release not published yet — using pinned 2024-25 edition")
        path = fetch(ERP_URL, "abs_erp_sa2.xlsx", max_age_days=180)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict] = {}
    year = None
    for sn in wb.sheetnames:
        if not sn.lower().startswith("table"):
            continue
        i_code = i_erp = None
        tbl_year = None
        prev: tuple = ()
        for r in wb[sn].iter_rows(min_row=1, max_row=None, values_only=True):
            if i_code is None:
                # The "ERP at 30 June" year columns sit in the row just above the
                # header row (other "no." columns are migration components — skip).
                if any(str(v).strip() == "SA2 code" for v in r if v):
                    i_code = next(i for i, v in enumerate(r) if str(v).strip() == "SA2 code")
                    ycols = {int(v): i for i, v in enumerate(prev)
                             if isinstance(v, (int, float)) and 2000 < v < 2100}
                    if not ycols:
                        break
                    tbl_year = max(ycols)
                    i_erp = ycols[tbl_year]
                    i_prev = ycols.get(tbl_year - 1)
                prev = r
                continue
            code = str(r[i_code] or "").strip()
            v = r[i_erp] if i_erp < len(r) else None
            pv = r[i_prev] if i_prev is not None and i_prev < len(r) else None
            if len(code) == 9 and code.isdigit() and isinstance(v, (int, float)):
                rec = {"erp": int(v), "erp_year": tbl_year}
                if isinstance(pv, (int, float)) and pv > 0:
                    rec["erp_growth_pct"] = round((v - pv) / pv * 100, 2)
                out[code] = rec
                year = tbl_year
    wb.close()
    cache.write_text(json.dumps(out), encoding="utf-8")
    print(f"  erp: {len(out)} SA2s (ERP at 30 June {year})")
    return out


if __name__ == "__main__":  # pragma: no cover
    e = get_erp()
    for code, nm in (("213051584", "Tarneit - North"), ("206061138", "Toorak"),
                     ("208031192", "Moorabbin Airport")):
        print(f"  {nm:20} {e.get(code)}")
