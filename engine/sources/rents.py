"""Rental medians + gross yield inputs per SA2 (Phase 4).

Source: Victorian DFFH quarterly "Rental Report" — Moving Annual Median Rent by
Suburb and Town workbook (free, official bond-lodgement data). Sheets per
dwelling type; columns are quarter (Count, Median) pairs back to Mar 2000.

We take, per suburb group (e.g. "Albert Park-Middle Park-West St Kilda"):
  * rent_weekly     — latest All-properties moving annual median
  * rent_12m        — % change vs 4 quarters earlier
  * house_rent      — latest 3-bedroom-house median (yield vs VG house median)
  * flat_rent       — latest 2-bedroom-flat median (yield vs VG unit median)
then attach to SA2s by locality-name matching (same approach as prices.py:
group names split into localities; an SA2 matches if any of its name tokens
equals a locality in the group).

The suburb groupings are a fixed list dating back to 2000, so newer growth
suburbs (Tarneit, Clyde North, Cobblebank...) are not named. For SA2s with no
suburb-group match we fall back to the companion "Quarterly Median Rents by
LGA" workbook (coverage flagged as "lga" for transparency).
"""
from __future__ import annotations

import re
import time

import openpyxl

from ..fetch import fetch, fetch_wayback

# Pinned fallback quarter (dffh.vic.gov.au slugs redirect to the xlsx asset).
# At build time we ask the data.vic CKAN catalogue for the NEWEST resource in
# each package so the data self-upgrades when DFFH publishes a quarter.
RENTS_URL = "https://www.dffh.vic.gov.au/moving-annual-rent-suburb-september-quarter-2025-excel"
RENTS_LGA_URL = "https://www.dffh.vic.gov.au/quarterly-median-rents-local-government-area-september-quarter-2025-excel"
RENTS_QUARTER = "Sep 2025"
_CKAN = "https://discover.data.vic.gov.au/api/3/action/package_show"
_PKG_SUBURB = "rental-report-quarterly-moving-annual-rents-by-suburb"
_PKG_LGA = "rental-report-quarterly-quarterly-median-rents-by-lga"


def _latest_resource(package: str, fallback: str) -> tuple[str, str]:
    """(url, label) of the newest resource in a CKAN package; pinned fallback."""
    import requests
    try:
        r = requests.get(_CKAN, params={"id": package},
                         headers={"User-Agent": "Mozilla/5.0 (melb-scorer)"}, timeout=45)
        res = r.json()["result"]["resources"]
        best = max(res, key=lambda x: x.get("created") or "")
        url = best.get("url") or fallback
        m = re.search(r"(march|june|september|december)-quarter-(\d{4})", url)
        label = f"{m.group(1)[:3].title()} {m.group(2)}" if m else RENTS_QUARTER
        return url, label
    except Exception as e:  # noqa: BLE001 - catalogue down -> pinned quarter
        print(f"  rents: CKAN lookup failed ({e}) — using pinned {RENTS_QUARTER}")
        return fallback, RENTS_QUARTER


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if s in ("", "-", "NA", "N/A", "n/a", "*"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _load_sheet(ws) -> dict[str, dict]:
    """{GROUP_NAME: {"latest": median, "prev4": median}} for one dwelling type."""
    rows = list(ws.iter_rows(values_only=True))
    # Row with quarter labels ("Mar 2000", ...) then a Count/Median row beneath.
    qrow = next(i for i, r in enumerate(rows[:6])
                if sum(1 for v in r if isinstance(v, str) and re.match(r"^(Mar|Jun|Sep|Dec) 20\d\d$", v)) >= 4)
    kind = rows[qrow + 1]
    med_cols = [ci for ci, v in enumerate(kind) if str(v).strip().lower() == "median"]
    if not med_cols:
        return {}
    latest_ci, prev_ci = med_cols[-1], (med_cols[-5] if len(med_cols) >= 5 else None)
    count_ci = latest_ci - 1     # Count column sits immediately left of each Median
    out = {}
    for r in rows[qrow + 2:]:
        name = r[1] if len(r) > 1 else None
        if not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        if name.lower().endswith("total") or name.lower() == "group":
            continue
        latest = _num(r[latest_ci]) if latest_ci < len(r) else None
        prev = _num(r[prev_ci]) if prev_ci is not None and prev_ci < len(r) else None
        bonds = _num(r[count_ci]) if 0 <= count_ci < len(r) else None
        if latest:
            out[name.upper()] = {"latest": latest, "prev4": prev, "bonds": bonds}
    return out


def _group_tokens(group: str) -> list[str]:
    """Localities in a DFFH group name: 'Albert Park-Middle Park-West St Kilda'."""
    return [t.strip().upper() for t in group.split("-") if t.strip()]


def _sa2_candidates(name: str) -> list[str]:
    n = re.sub(r"\(.*?\)", "", name)
    parts = [p.strip().upper() for p in re.split(r"\s*-\s*", n) if p.strip()]
    whole = re.sub(r"\s+", " ", n).strip().upper()
    out = []
    for c in parts + [whole]:
        if c and c not in out:
            out.append(c)
    return out


def _sheet(wb, *needles):
    """Find a sheet whose name contains all needles (case/space-insensitive)."""
    for sn in wb.sheetnames:
        low = sn.lower().replace(" ", "")
        if all(n in low for n in needles):
            return wb[sn]
    raise KeyError(needles)


def _load_workbook_sheets(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    allp = _load_sheet(_sheet(wb, "allpropert"))
    house = _load_sheet(_sheet(wb, "3b", "house"))
    flat = _load_sheet(_sheet(wb, "2b", "flat"))
    wb.close()
    return allp, house, flat


def _avg(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else None


def get_rents(name_by_code: dict[str, str],
              lga_by_code: dict[str, str] | None = None) -> dict[str, dict]:
    """{sa2_code: {rent_weekly, rent_12m, house_rent, flat_rent, rent_bonds,
                   rent_quarter, rent_source}}"""
    def _dffh(url: str, fname: str):
        # dffh.vic.gov.au is very slow / bot-hostile from datacenter IPs — fall
        # back to the Wayback Machine if the direct download keeps timing out.
        try:
            return fetch(url, fname, max_age_days=60)
        except Exception as e:  # noqa: BLE001
            print(f"  direct DFFH download failed ({e}); trying Wayback ...")
            return fetch_wayback(url, fname)

    sub_url, quarter = _latest_resource(_PKG_SUBURB, RENTS_URL)
    lga_url, _ = _latest_resource(_PKG_LGA, RENTS_LGA_URL)
    try:
        allp, house, flat = _load_workbook_sheets(_dffh(sub_url, "dffh_rents_by_suburb.xlsx"))
    except Exception as e:  # noqa: BLE001 - the workflow's coverage guard stops a bad commit
        print(f"  suburb rents workbook unavailable ({e}); shipping no rents this build")
        allp, house, flat = {}, {}, {}
    # The LGA workbook only backfills the handful of growth suburbs missing from
    # the suburb list — degrade gracefully rather than failing the whole build
    # (DFFH throttles back-to-back downloads from one IP).
    try:
        time.sleep(20)
        lga_allp, lga_house, lga_flat = _load_workbook_sheets(_dffh(lga_url, "dffh_rents_by_lga.xlsx"))
    except Exception as e:  # noqa: BLE001
        print(f"  LGA rents workbook unavailable ({e}); using suburb-level rents only")
        lga_allp, lga_house, lga_flat = {}, {}, {}

    # locality -> group name (a locality can appear in exactly one group)
    loc2group: dict[str, str] = {}
    for group in allp:
        for tok in _group_tokens(group):
            loc2group.setdefault(tok, group)

    out = {}
    for code, name in name_by_code.items():
        groups = []
        for cand in _sa2_candidates(name):
            g = loc2group.get(cand)
            if g and g not in groups:
                groups.append(g)
        rw = _avg([allp[g]["latest"] for g in groups if g in allp])
        prev = _avg([allp[g]["prev4"] for g in groups if g in allp and allp[g]["prev4"]])
        rec = {
            "rent_weekly": rw,
            "rent_12m": round((rw - prev) / prev * 100, 1) if rw and prev else None,
            "house_rent": _avg([house[g]["latest"] for g in groups if g in house]),
            "flat_rent": _avg([flat[g]["latest"] for g in groups if g in flat]),
            # annual bond lodgements = rental-market liquidity (thin markets rank
            # alongside liquid ones otherwise)
            "rent_bonds": _avg([allp[g]["bonds"] for g in groups if g in allp and allp[g].get("bonds")]),
            "rent_source": "suburb" if rw else None,
        }
        # LGA fallback for the (post-2000) growth suburbs the DFFH list never named.
        if rec["rent_weekly"] is None and lga_by_code:
            lga = (lga_by_code.get(code) or "").strip().upper()
            lga = {"MERRI-BEK": "MORELAND"}.get(lga, lga)   # DFFH kept the old LGA name
            L = lga_allp.get(lga)
            if L:
                rec["rent_weekly"] = L["latest"]
                rec["rent_12m"] = (round((L["latest"] - L["prev4"]) / L["prev4"] * 100, 1)
                                   if L["prev4"] else None)
                rec["house_rent"] = (lga_house.get(lga) or {}).get("latest")
                rec["flat_rent"] = (lga_flat.get(lga) or {}).get("latest")
                rec["rent_source"] = "lga"
        rec["rent_quarter"] = quarter if rec["rent_weekly"] else None
        out[code] = rec
    cov = sum(1 for v in out.values() if v["rent_weekly"])
    sub = sum(1 for v in out.values() if v["rent_source"] == "suburb")
    print(f"  rents: {cov}/{len(out)} SA2s covered ({sub} suburb-level, "
          f"{cov - sub} LGA fallback; {quarter})")
    return out


if __name__ == "__main__":  # pragma: no cover
    import json
    from .. import config
    fc = json.loads((config.CITY_DATA / config.BOUNDARIES_NAME).read_text(encoding="utf-8"))
    names = {f["properties"]["sa2_code"]: f["properties"]["sa2_name"] for f in fc["features"]}
    from .crime import get_crime
    lgas = {c: v["lga"] for c, v in get_crime().items()}
    r = get_rents(names, lgas)
    for nm in ("Toorak", "Tarneit - North", "Brighton (Vic.)", "Nunawading", "Carlton"):
        code = next((c for c, n in names.items() if n == nm), None)
        if code:
            print(f"  {nm:24} {r[code]}")
