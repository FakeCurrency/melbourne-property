"""Crime inputs per SA2, split by offence type so personal safety can be
prioritised over property crime in the Liveability score.

Two resolutions, merged in build.py (suburb preferred, LGA fallback):

LGA level — CSA "Recorded Offences by LGA" workbook:
  * Table 01 -> each LGA's total offence count + rate  => implied population
  * Table 02 -> offence counts by Offence Division per LGA
  attached to each SA2 by point-in-polygon against ABS LGA boundaries.

Suburb level (Phase 4) — CSA "Criminal Incidents by LGA" workbook, Table 03:
  criminal incidents by suburb/town × offence division. Suburb counts are
  allocated to SA2s by locality-name matching (as prices.py) with the counts
  split across matching SA2s in proportion to population, then divided by SA2
  population for a per-100k rate. This fixes the old LGA artifact where e.g.
  Toorak wore Stonnington-wide (Chapel St precinct) rates.

Both compute, per 100k population:
  person   = Division "A Crimes against the person"
  property = Division "B Property and deception offences"
  total    = all divisions
"""
from __future__ import annotations

import json
import re
import zipfile
from collections import defaultdict

import openpyxl
import shapefile

from .. import config, geo
from ..fetch import fetch

# Pinned fallback release (year ending March 2026). CSA publishes quarterly on
# a fixed URL pattern — _discover_csa() probes newer quarters first so the data
# self-upgrades when a release lands.
_CSA_BASE = "https://files.crimestatistics.vic.gov.au"
_CSA_PIN = ("2026-06", "March", 2026)   # (folder, quarter-ending month, year)
_CSA_SEQ = ["March", "June", "September", "December"]


def _csa_candidates():
    """Newest-first candidate (folder, month, year) tuples from the pinned one."""
    folder, month, year = _CSA_PIN
    i = _CSA_SEQ.index(month)
    fy, fm = int(folder.split("-")[0]), int(folder.split("-")[1])
    cands = []
    for step in (4, 3, 2, 1, 0):        # up to 4 quarters ahead of the pin
        qi = (i + step) % 4
        yr = year + (i + step) // 4
        months_ahead = 3 * step
        f_y, f_m = fy + (fm - 1 + months_ahead) // 12, (fm - 1 + months_ahead) % 12 + 1
        cands.append((f"{f_y}-{f_m:02d}", _CSA_SEQ[qi], yr))
    return cands


def _discover_csa(kind: str, fname: str):
    """Fetch the newest available CSA workbook of ``kind``; pinned fallback."""
    from ..fetch import fresh
    dest = config.DATA_RAW / fname
    if fresh(dest, 60):
        print(f"  cached  {fname}")
        return dest
    for folder, month, year in _csa_candidates():
        url = f"{_CSA_BASE}/{folder}/Data_Tables_LGA_{kind}_Year_Ending_{month}_{year}.xlsx"
        try:
            return fetch(url, fname, force=True)
        except Exception:
            continue
    raise RuntimeError(f"No CSA {kind} workbook reachable (tried 5 quarters)")
LGA_SHP_URL = (
    "https://www.abs.gov.au/statistics/standards/"
    "australian-statistical-geography-standard-asgs/edition-3-july-2021-june-2026/"
    "access-and-downloads/digital-boundary-files/LGA_2022_AUST_GDA2020_SHP.zip"
)

# ABS boundary names that CSA publishes under a newer name (LGA renames).
_ALIASES = {"moreland": "merri-bek"}


def _norm_lga(name: str) -> str:
    s = str(name).lower()
    for junk in ("(vic.)", "(c)", "(rc)", "(s)", "(b)", "(m)"):
        s = s.replace(junk, "")
    s = s.strip()
    return _ALIASES.get(s, s)


def _offences_by_lga() -> dict[str, dict]:
    """{lga_key: {person, property, total}} rates per 100k for the latest year."""
    from ..fetch import fresh
    cache = config.DATA_RAW / "crime_offences_by_lga.json"
    if fresh(cache, 60):
        print("  cached  crime_offences_by_lga.json")
        return json.loads(cache.read_text(encoding="utf-8"))

    path = _discover_csa("Recorded_Offences", "csa_lga_recorded_offences.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Table 01: LGA totals -> implied population = count / rate * 100k
    pop, total_rate = {}, {}
    years = set()
    for year, _end, _region, lga, count, rate in wb["Table 01"].iter_rows(min_row=2, values_only=True):
        if year is None or lga is None or not rate or _norm_lga(lga) == "total":
            continue
        years.add(int(year))
    latest = max(years)
    for year, _end, _region, lga, count, rate in wb["Table 01"].iter_rows(min_row=2, values_only=True):
        if int(year or 0) != latest or lga is None or not rate or not count:
            continue
        key = _norm_lga(lga)
        if key == "total":
            continue
        pop[key] = count / float(rate) * 1e5
        total_rate[key] = float(rate)

    # Table 02: offence counts by Offence Division per LGA
    person, prop, total = defaultdict(float), defaultdict(float), defaultdict(float)
    for row in wb["Table 02"].iter_rows(min_row=2, values_only=True):
        year, _end, _psa, lga, div, _sub, _grp, count = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        if int(year or 0) != latest or lga is None or not count:
            continue
        key = _norm_lga(lga)
        total[key] += count
        if str(div).startswith("A "):        # crimes against the person
            person[key] += count
        elif str(div).startswith("B "):      # property and deception offences
            prop[key] += count
    wb.close()

    out = {}
    for key, p in pop.items():
        if p <= 0:
            continue
        out[key] = {
            "person": round(person[key] / p * 1e5, 1),
            "property": round(prop[key] / p * 1e5, 1),
            "total": round(total[key] / p * 1e5, 1) or total_rate.get(key),
        }
    cache.write_text(json.dumps(out), encoding="utf-8")
    print(f"  crime: {len(out)} LGAs, year {latest} (person/property/total rates)")
    return out


def _load_vic_lgas() -> list[dict]:
    """Victorian LGA polygons with bounding boxes for fast point-in-polygon."""
    zip_path = fetch(LGA_SHP_URL, "LGA_2022_SHP.zip")
    ed = config.DATA_RAW / "lga_shp"
    if not ed.exists():
        with zipfile.ZipFile(zip_path) as z:
            z.extractall(ed)
    shp = next(ed.glob("*.shp"))
    r = shapefile.Reader(str(shp))
    fields = [f[0] for f in r.fields[1:]]
    i_name = next(i for i, f in enumerate(fields) if f.upper().startswith("LGA_NAME"))
    i_ste = next(i for i, f in enumerate(fields) if f.upper().startswith("STE_NAME"))
    lgas = []
    for sr in r.iterShapeRecords():
        if str(sr.record[i_ste]).strip() != config.STATE_NAME or not sr.shape.points:
            continue
        geom = sr.shape.__geo_interface__
        polys = [geom["coordinates"]] if geom["type"] == "Polygon" else geom["coordinates"]
        rings = [[tuple(c) for c in poly[0]] for poly in polys]
        xs = [x for ring in rings for x, _ in ring]
        ys = [y for ring in rings for _, y in ring]
        lgas.append({
            "name": str(sr.record[i_name]), "key": _norm_lga(sr.record[i_name]),
            "rings": rings, "bbox": (min(xs), min(ys), max(xs), max(ys)),
        })
    return lgas


def get_crime() -> dict[str, dict]:
    """{sa2_code: {lga, person, property, total}} for Greater Melbourne SA2s."""
    rates = _offences_by_lga()
    lgas = _load_vic_lgas()
    points = geo.sa2_points()

    out = {}
    for code, (x, y) in points.items():
        chosen = None
        for lga in lgas:
            x0, y0, x1, y1 = lga["bbox"]
            if x0 <= x <= x1 and y0 <= y <= y1 and any(geo.point_in_ring(x, y, r) for r in lga["rings"]):
                chosen = lga
                break
        if chosen is None:
            chosen = min(lgas, key=lambda L: (
                (x - (L["bbox"][0] + L["bbox"][2]) / 2) ** 2
                + (y - (L["bbox"][1] + L["bbox"][3]) / 2) ** 2))
        r = rates.get(chosen["key"], {})
        out[code] = {
            "lga": chosen["name"],
            "person": r.get("person"), "property": r.get("property"), "total": r.get("total"),
        }
    matched = sum(1 for v in out.values() if v["person"] is not None)
    print(f"  crime: matched {matched}/{len(out)} SA2s to an LGA rate")
    return out


# --- suburb-level (Phase 4) -------------------------------------------------
_COMPASS = {"EAST", "WEST", "NORTH", "SOUTH", "CENTRAL", "INNER", "OUTER"}


def _sa2_localities(name: str) -> list[str]:
    """Localities an SA2 name refers to (UPPERCASE), for suburb-count matching."""
    n = re.sub(r"\(.*?\)", "", name)
    parts = [p.strip().upper() for p in re.split(r"\s*-\s*", n) if p.strip()]
    out = []
    for p in parts + [re.sub(r"\s+", " ", n).strip().upper()]:
        if p and p not in out and p not in _COMPASS:
            out.append(p)
        # "Melbourne CBD - East" should also match the CSA locality "Melbourne"
        if p.endswith(" CBD") and p[:-4] not in out:
            out.append(p[:-4])
    return out


def _incidents_by_suburb(metro_lgas: set[str]) -> dict[str, dict]:
    """{SUBURB: {person, property, total, person_prev}} incident counts —
    latest year plus person counts four years earlier (for the trend arrow)."""
    from ..fetch import fresh
    cache = config.DATA_RAW / "crime_incidents_by_suburb_v2.json"
    if fresh(cache, 60):
        print("  cached  crime_incidents_by_suburb_v2.json")
        return json.loads(cache.read_text(encoding="utf-8"))

    path = _discover_csa("Criminal_Incidents", "csa_lga_criminal_incidents.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 03"]
    latest = 0
    rows = []
    for year, _end, lga, _pc, suburb, div, _sub, _grp, count in ws.iter_rows(min_row=2, values_only=True):
        if year is None or suburb is None or not count:
            continue
        y = int(year)
        latest = max(latest, y)
        rows.append((y, _norm_lga(lga), str(suburb).strip().upper(), str(div), float(count)))
    wb.close()

    prev_year = latest - 4
    counts: dict[str, dict] = defaultdict(
        lambda: {"person": 0.0, "property": 0.0, "total": 0.0, "person_prev": 0.0})
    for y, lga, suburb, div, count in rows:
        if lga not in metro_lgas:
            continue
        c = counts[suburb]
        if y == latest:
            c["total"] += count
            if div.startswith("A "):
                c["person"] += count
            elif div.startswith("B "):
                c["property"] += count
        elif y == prev_year and div.startswith("A "):
            c["person_prev"] += count
    out = dict(counts)
    cache.write_text(json.dumps(out), encoding="utf-8")
    print(f"  crime: suburb incident counts for {len(out)} metro localities, "
          f"year {latest} (trend vs {prev_year})")
    return out


def get_crime_suburb(name_by_code: dict[str, str],
                     pop_by_code: dict[str, float],
                     lga_by_code: dict[str, str]) -> dict[str, dict]:
    """{sa2_code: {person, property, total}} per-100k rates from suburb incidents.

    Locality counts are split across the SA2s whose names reference the locality
    (population-weighted), then summed per SA2 and divided by SA2 population.
    """
    metro_lgas = {_norm_lga(v) for v in lga_by_code.values() if v}
    counts = _incidents_by_suburb(metro_lgas)

    loc2sa2: dict[str, list[str]] = defaultdict(list)
    for code, name in name_by_code.items():
        if not pop_by_code.get(code):
            continue
        for loc in _sa2_localities(name):
            loc2sa2[loc].append(code)

    alloc: dict[str, dict] = defaultdict(
        lambda: {"person": 0.0, "property": 0.0, "total": 0.0, "person_prev": 0.0})
    for loc, c in counts.items():
        sa2s = loc2sa2.get(loc)
        if not sa2s:
            continue
        pop_sum = sum(pop_by_code[s] for s in sa2s)
        if pop_sum <= 0:
            continue
        for s in sa2s:
            w = pop_by_code[s] / pop_sum
            a = alloc[s]
            for k in ("person", "property", "total", "person_prev"):
                a[k] += c.get(k, 0.0) * w

    out = {}
    for code in name_by_code:
        pop = pop_by_code.get(code)
        a = alloc.get(code)
        if not pop or not a or a["total"] <= 0:
            out[code] = {}
            continue
        rec = {k: round(v / pop * 1e5, 1) for k, v in a.items() if k != "person_prev"}
        # personal-crime direction over ~4 years (same denominator both sides,
        # so population drift cancels out of the ratio)
        if a["person_prev"] >= 20 and a["person"] >= 20:
            rec["person_trend_pct"] = round((a["person"] - a["person_prev"])
                                            / a["person_prev"] * 100, 1)
        out[code] = rec
    matched = sum(1 for v in out.values() if v)
    print(f"  crime: suburb-level rates for {matched}/{len(out)} SA2s (rest fall back to LGA)")
    return out


def get_postcode_map(name_by_code: dict[str, str],
                     lga_by_code: dict[str, str]) -> dict[str, list[str]]:
    """{postcode: [sa2_codes]} for search — from CSA Table 03's postcode column."""
    from ..fetch import fresh
    cache = config.DATA_RAW / "crime_postcodes.json"
    if fresh(cache, 180):
        print("  cached  crime_postcodes.json")
        pc2loc = json.loads(cache.read_text(encoding="utf-8"))
    else:
        metro_lgas = {_norm_lga(v) for v in lga_by_code.values() if v}
        path = _discover_csa("Criminal_Incidents", "csa_lga_criminal_incidents.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        pairs = set()
        for _y, _end, lga, pc, suburb, *_rest in wb["Table 03"].iter_rows(min_row=2, values_only=True):
            if pc is None or suburb is None or _norm_lga(lga) not in metro_lgas:
                continue
            pc = str(pc).strip().split(".")[0]
            if len(pc) == 4 and pc.isdigit():
                pairs.add((pc, str(suburb).strip().upper()))
        wb.close()
        pc2loc = defaultdict(list)
        for pc, loc in sorted(pairs):
            pc2loc[pc].append(loc)
        pc2loc = dict(pc2loc)
        cache.write_text(json.dumps(pc2loc), encoding="utf-8")
        print(f"  postcodes: {len(pc2loc)} metro postcodes mapped to localities")

    loc2sa2: dict[str, list[str]] = defaultdict(list)
    for code, name in name_by_code.items():
        for loc in _sa2_localities(name):
            loc2sa2[loc].append(code)
    out: dict[str, list[str]] = {}
    for pc, locs in pc2loc.items():
        codes: list[str] = []
        for loc in locs:
            for c in loc2sa2.get(loc, []):
                if c not in codes:
                    codes.append(c)
        if codes:
            out[pc] = codes
    print(f"  postcodes: {len(out)} postcodes resolve to SA2s")
    return out


if __name__ == "__main__":  # pragma: no cover
    c = get_crime()
    for code in list(c)[:3]:
        print(code, c[code])
